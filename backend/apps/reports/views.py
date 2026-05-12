import os
import io
import zipfile
import hashlib
import datetime
from collections import defaultdict

from django.http import HttpResponse
from django.conf import settings
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status

from apps.users.permissions import IsExamDept
from apps.evaluation.models import EvaluationResult
from apps.scanning.models import Subject, Bundle, AnswerSheet
from utils.audit_helper import log_action
from apps.evaluation.result_page_generator import generate_result_page

from pypdf import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .excel_export import generate_excel_report
from .pdf_export import generate_student_pdf_response, generate_student_pdf_bytes, generate_bundle_pdf_response


# ── Helper ────────────────────────────────────────
def _build_student_data(request):
    """
    Query EvaluationResults filtered by request query params and return:
    - students: list of grouped student dicts
    - filters: dict of available filter values for dropdowns
    """
    department = request.query_params.get('department', '')
    semester = request.query_params.get('semester', '')
    academic_year = request.query_params.get('academic_year', '')
    subject_id = request.query_params.get('subject', '')
    roll_number = request.query_params.get('roll_number', '')

    qs = EvaluationResult.objects.select_related(
        'answer_sheet',
        'answer_sheet__bundle',
        'answer_sheet__bundle__subject',
        'teacher',
    ).filter(answer_sheet__bundle__status='submitted')

    if department:
        qs = qs.filter(answer_sheet__bundle__subject__department=department)
    if semester:
        qs = qs.filter(answer_sheet__bundle__subject__semester=int(semester))
    if academic_year:
        qs = qs.filter(answer_sheet__bundle__academic_year=academic_year)
    if subject_id:
        qs = qs.filter(answer_sheet__bundle__subject_id=subject_id)
    if roll_number:
        qs = qs.filter(answer_sheet__roll_number__icontains=roll_number)

    # Group by roll_number
    grouped = defaultdict(lambda: {
        'roll_number': '',
        'department': '',
        'semester': '',
        'academic_year': '',
        'subjects': [],
        'grand_total': 0,
    })

    for result in qs:
        sheet = result.answer_sheet
        bundle = sheet.bundle
        subject = bundle.subject
        roll = sheet.roll_number

        entry = grouped[roll]
        entry['roll_number'] = roll
        entry['department'] = subject.department
        entry['semester'] = subject.semester
        entry['academic_year'] = bundle.academic_year

        # Get max_marks from marking scheme if available
        max_marks = 0
        try:
            max_marks = subject.marking_scheme.total_marks
        except Exception:
            pass

        entry['subjects'].append({
            'subject_code': subject.subject_code,
            'subject_name': subject.subject_name,
            'semester': subject.semester,
            'total_marks': result.total_marks,
            'max_marks': max_marks,
            'evaluated_on': result.graded_at.strftime('%d/%m/%Y') if result.graded_at else '',
        })

    # Compute grand totals
    students = []
    for roll, data in sorted(grouped.items()):
        data['grand_total'] = sum(s['total_marks'] for s in data['subjects'])
        students.append(data)

    # Build available filter values
    all_bundles = Bundle.objects.filter(status='submitted').select_related('subject')
    departments_list = sorted(set(b.subject.department for b in all_bundles if b.subject.department))
    semesters_list = sorted(set(b.subject.semester for b in all_bundles))
    academic_years_list = sorted(set(b.academic_year for b in all_bundles if b.academic_year), reverse=True)
    subjects_list = [
        {'id': s.id, 'code': s.subject_code, 'name': s.subject_name}
        for s in Subject.objects.filter(
            bundles__status='submitted'
        ).distinct().order_by('subject_code')
    ]

    filters = {
        'departments': departments_list,
        'semesters': semesters_list,
        'academic_years': academic_years_list,
        'subjects': subjects_list,
    }

    return students, filters


# ── Views ─────────────────────────────────────────

class ReportsSummaryView(APIView):
    """
    GET /api/reports/
    Returns student-grouped evaluation results with filter values.
    Query params: department, semester, academic_year, subject, roll_number
    """
    permission_classes = [IsExamDept]

    def get(self, request):
        students, filters = _build_student_data(request)
        return Response({
            'students': students,
            'filters': filters,
            'count': len(students),
        })


class ExcelExportView(APIView):
    """GET /api/reports/export/excel/ — Download grouped results as Excel."""
    permission_classes = [IsExamDept]

    def get(self, request):
        students, _ = _build_student_data(request)

        log_action(
            request, 'RESULT_GENERATED', 'Report', 0,
            notes=f'Excel report exported. {len(students)} students.'
        )

        return generate_excel_report(students)


class StudentPDFExportView(APIView):
    """GET /api/reports/export/student-pdf/<roll_number>/ — Single student PDF."""
    permission_classes = [IsExamDept]

    def get(self, request, roll_number):
        # Build full student data then find the specific student
        students, _ = _build_student_data(request)
        student = next((s for s in students if s['roll_number'] == roll_number), None)

        if not student:
            return Response(
                {'error': f'No results found for roll number {roll_number}.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        log_action(
            request, 'RESULT_GENERATED', 'Report', 0,
            notes=f'Student PDF exported for roll {roll_number}.'
        )

        return generate_student_pdf_response(student)


class AllPDFsExportView(APIView):
    """GET /api/reports/export/all-pdfs/ — ZIP of all student PDFs matching filters."""
    permission_classes = [IsExamDept]

    def get(self, request):
        students, _ = _build_student_data(request)

        if not students:
            return Response(
                {'error': 'No results found for the current filters.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Build ZIP in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for student in students:
                roll = student['roll_number']
                pdf_bytes = generate_student_pdf_bytes(student)
                zf.writestr(f'result_{roll}.pdf', pdf_bytes.read())

        zip_buffer.seek(0)

        log_action(
            request, 'RESULT_GENERATED', 'Report', 0,
            notes=f'Bulk PDF ZIP exported. {len(students)} students.'
        )

        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="all_student_results.zip"'
        return response


class BundlePDFExportView(APIView):
    """
    GET /api/reports/export/bundle-pdf/<bundle_id>/
    Download a single PDF that covers every answer sheet in a bundle.
    Each sheet gets its own page: roll number, token/code, and a
    section-by-section / question-by-question marks table.
    """
    permission_classes = [IsExamDept]

    def get(self, request, bundle_id):
        from datetime import date

        try:
            bundle = Bundle.objects.select_related('subject').get(pk=bundle_id)
        except Bundle.DoesNotExist:
            return Response({'error': 'Bundle not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Fetch all answer sheets for this bundle with their evaluations
        # EvaluationResult is a ForeignKey (related_name='evaluations'), not OneToOne
        sheets_qs = AnswerSheet.objects.filter(bundle=bundle).prefetch_related(
            'evaluations'
        ).order_by('roll_number')

        sheets_data = []
        for sheet in sheets_qs:
            # Pick the best evaluation: is_final first, then assessor, then latest
            evals = list(sheet.evaluations.all())
            evaluation = None
            if evals:
                final_evals = [e for e in evals if e.is_final]
                assessor_evals = [e for e in evals if e.role == 'assessor']
                if final_evals:
                    evaluation = final_evals[0]
                elif assessor_evals:
                    evaluation = assessor_evals[0]
                else:
                    evaluation = evals[0]

            sheets_data.append({
                'roll_number': sheet.roll_number or 'N/A',
                'token': sheet.token or 'N/A',
                'status': sheet.status,
                'total_marks': evaluation.total_marks if evaluation else '-',
                'section_results': evaluation.section_results if evaluation else [],
            })

        bundle_data = {
            'bundle_number': bundle.bundle_number,
            'subject_code': bundle.subject.subject_code,
            'subject_name': bundle.subject.subject_name,
            'department': bundle.subject.department,
            'semester': bundle.subject.semester,
            'academic_year': bundle.academic_year,
            'total_sheets': bundle.total_sheets,
            'generated_on': date.today().strftime('%d %B %Y'),
            'sheets': sheets_data,
        }

        log_action(
            request, 'RESULT_GENERATED', 'Bundle', bundle_id,
            notes=f'Bundle PDF exported for bundle #{bundle.bundle_number} ({len(sheets_data)} sheets).'
        )

        return generate_bundle_pdf_response(bundle_data)


def _prepend_result_page(result_page_bytes: bytes, marked_pdf_path: str) -> bytes:
    """
    Prepends the result summary page to the marked PDF.
    Everything happens in memory — nothing written to disk.

    Args:
        result_page_bytes: bytes of the result page PDF (from generate_result_page)
        marked_pdf_path:   absolute path to the marked PDF on disk

    Returns:
        bytes of the combined PDF (result page + all marked pages)
    """
    writer = PdfWriter()

    # Page 1: result summary
    result_reader = PdfReader(io.BytesIO(result_page_bytes))
    writer.add_page(result_reader.pages[0])

    # Pages 2+: original marked PDF pages
    marked_reader = PdfReader(marked_pdf_path)
    for page in marked_reader.pages:
        writer.add_page(page)

    output = io.BytesIO()
    writer.write(output)
    return output.getvalue()


def _build_summary_xlsx(results_data: list, subject_code: str, subject_name: str) -> bytes:
    """
    Builds a summary Excel file listing all students with their marks.

    Args:
        results_data: list of dicts — each has:
            { roll_number, token, total_marks, max_marks, section_results }
        subject_code: e.g. "ETH405"
        subject_name: e.g. "Embedded Systems"

    Returns:
        bytes of the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{subject_code} Results"

    # Styles
    header_fill   = PatternFill("solid", fgColor="0F6E56")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    title_font    = Font(bold=True, size=12)
    total_fill    = PatternFill("solid", fgColor="E1F5EE")
    total_font    = Font(bold=True, color="0F6E56", size=10)
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center")
    thin_border   = Border(
        left=Side(style="thin", color="CBD5E0"),
        right=Side(style="thin", color="CBD5E0"),
        top=Side(style="thin", color="CBD5E0"),
        bottom=Side(style="thin", color="CBD5E0"),
    )

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Results Summary — {subject_code}: {subject_name}"
    ws["A1"].font      = title_font
    ws["A1"].alignment = left_align
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font      = Font(size=9, color="4A5568")
    ws["A2"].alignment = left_align
    ws.row_dimensions[2].height = 16

    # Header row
    headers = ["#", "Roll Number", "Token / Code", "Max Marks", "Marks Obtained", "Percentage"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[3].height = 18

    # Data rows
    for row_idx, data in enumerate(results_data, start=1):
        row_num   = row_idx + 3
        max_marks = data.get('max_marks', 0)
        obtained  = data.get('total_marks', 0)
        pct       = round((obtained / max_marks * 100), 1) if max_marks > 0 else 0

        row_values = [
            row_idx,
            data['roll_number'],
            data['token'],
            max_marks,
            obtained,
            f"{pct}%",
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = center_align if col_idx != 2 else left_align
            cell.border    = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9FAFB")

        ws.row_dimensions[row_num].height = 16

    # Totals row
    total_row = len(results_data) + 4
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws[f"A{total_row}"] = f"Total Students: {len(results_data)}"
    ws[f"A{total_row}"].font      = total_font
    ws[f"A{total_row}"].fill      = total_fill
    ws[f"A{total_row}"].alignment = left_align
    ws[f"A{total_row}"].border    = thin_border

    avg_obtained = round(
        sum(d.get('total_marks', 0) for d in results_data) / len(results_data), 1
    ) if results_data else 0
    ws[f"E{total_row}"] = f"Avg: {avg_obtained}"
    ws[f"E{total_row}"].font      = total_font
    ws[f"E{total_row}"].fill      = total_fill
    ws[f"E{total_row}"].alignment = center_align
    ws[f"E{total_row}"].border    = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class DownloadMarkedPDFsView(APIView):
    """
    GET /api/reports/download-marked-pdfs/
    exam_dept only.

    Query params (use one):
        subject_id  — all completed sheets across all bundles for this subject
        bundle_id   — all completed sheets in this specific bundle

    Returns a ZIP file containing:
        marked_papers_{subject_code}.zip
          ├── {roll_number}_{subject_code}_marked.pdf   ← one per student
          ├── {roll_number}_{subject_code}_marked.pdf
          ├── ...
          ├── results_summary_{subject_code}.xlsx
          └── manifest.txt

    Each PDF = result summary page (A4, auto-generated) + original marked pages.
    Files on disk are NEVER modified — all merging is done in memory.
    Roll numbers appear in ZIP filenames but never on disk storage.
    """
    permission_classes = [IsExamDept]

    def get(self, request):
        subject_id = request.query_params.get('subject_id')
        bundle_id  = request.query_params.get('bundle_id')

        if not subject_id and not bundle_id:
            return Response(
                {'error': 'Provide either subject_id or bundle_id.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Fetch completed evaluations ───────────────────────────────────────
        qs = EvaluationResult.objects.select_related(
            'answer_sheet',
            'answer_sheet__bundle',
            'answer_sheet__bundle__subject',
        ).exclude(
            marked_pdf_path__isnull=True
        ).exclude(
            marked_pdf_path=''
        ).filter(
            answer_sheet__status='completed'
        )

        if bundle_id:
            qs = qs.filter(answer_sheet__bundle_id=bundle_id)
        elif subject_id:
            qs = qs.filter(answer_sheet__bundle__subject_id=subject_id)

        qs = qs.order_by('answer_sheet__roll_number')

        if not qs.exists():
            return Response(
                {'error': 'No completed marked PDFs found for the given filter.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # ── Gather metadata ───────────────────────────────────────────────────
        first         = qs.first()
        subject       = first.answer_sheet.bundle.subject
        subject_code  = subject.subject_code
        subject_name  = subject.subject_name
        department    = subject.department
        semester      = subject.semester
        total_count   = qs.count()

        # Infer academic year from current date
        now = datetime.datetime.now()
        if now.month >= 6:
            academic_year = f"{now.year}-{str(now.year + 1)[2:]}"
        else:
            academic_year = f"{now.year - 1}-{str(now.year)[2:]}"

        # ── Build ZIP in memory ───────────────────────────────────────────────
        zip_buffer    = io.BytesIO()
        missing_files = []
        included      = []
        results_data  = []   # for the summary Excel

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:

            for idx, result in enumerate(qs, start=1):
                sheet       = result.answer_sheet
                roll_number = sheet.roll_number
                token       = sheet.token
                full_path   = os.path.join(settings.MEDIA_ROOT, result.marked_pdf_path)

                # Skip if file missing on disk
                if not os.path.exists(full_path):
                    missing_files.append(
                        f"{roll_number} ({token}) — marked PDF missing on disk: "
                        f"{result.marked_pdf_path}"
                    )
                    continue

                # Compute max marks from section_results
                max_marks = sum(
                    q.get('max_marks', 0)
                    for sec in result.section_results
                    for q in sec.get('questions', [])
                )

                # ── Generate result summary page ──────────────────────────────
                try:
                    result_page_bytes = generate_result_page(
                        roll_number=roll_number,
                        token=token,
                        subject_code=subject_code,
                        subject_name=subject_name,
                        department=department,
                        semester=semester,
                        academic_year=academic_year,
                        section_results=result.section_results,
                        total_marks=result.total_marks,
                        student_index=idx,
                        total_students=total_count,
                    )
                except Exception as e:
                    missing_files.append(
                        f"{roll_number} — result page generation failed: {e}"
                    )
                    continue

                # ── Prepend result page to marked PDF (in memory) ─────────────
                try:
                    combined_pdf_bytes = _prepend_result_page(
                        result_page_bytes=result_page_bytes,
                        marked_pdf_path=full_path,
                    )
                except Exception as e:
                    missing_files.append(
                        f"{roll_number} — PDF merge failed: {e}"
                    )
                    continue

                # ── Add to ZIP with roll-number-based filename ────────────────
                zip_filename = f"{roll_number}_{subject_code}_marked.pdf"
                zf.writestr(zip_filename, combined_pdf_bytes)
                included.append(roll_number)

                # Collect for summary Excel
                results_data.append({
                    'roll_number': roll_number,
                    'token':       token,
                    'total_marks': result.total_marks,
                    'max_marks':   max_marks,
                    'section_results': result.section_results,
                })

            # Removed Summary Excel and Manifest generation as requested by user

        # ── AuditLog ──────────────────────────────────────────────────────────
        log_action(
            request,
            action_type='RESULT_GENERATED',
            target_model='Subject',
            target_id=int(subject_id) if subject_id else 0,
            new_value={
                'included_count': len(included),
                'missing_count':  len(missing_files),
                'subject_code':   subject_code,
                'filter':         f"subject_id={subject_id}" if subject_id
                                  else f"bundle_id={bundle_id}",
            },
            notes=f"Bulk marked PDF download by {request.user.full_name}"
        )

        # ── Stream ZIP response ───────────────────────────────────────────────
        zip_buffer.seek(0)
        zip_filename = f"marked_papers_{subject_code}.zip"
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        return response
